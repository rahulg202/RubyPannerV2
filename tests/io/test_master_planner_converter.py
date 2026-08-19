"""Tests for the Master Planner converter (io_adapters.master_planner_converter).

Builds a synthetic workbook shaped like the real Schedule sheet — header on row 2,
a spacer row above, aggregate columns, EU shading on selected headers — so these
tests never depend on customer data.
"""

from __future__ import annotations

import io
from datetime import datetime

import openpyxl
import pytest
from openpyxl.styles import PatternFill

from domain.site_derivation import mapping_frame, notes_frame, sites_frame
from io_adapters.master_planner_converter import (
    EU_RESTRICTED_FILL,
    MasterPlannerConverter,
    convert_master_planner,
)
from io_adapters.master_planner_parser import assign_stable_id, parse_master_planner

# Customer columns start after the ten aggregate columns and one marker column.
CUSTOMER_HEADERS = [
    "00449    Alpha Specialty Care, Fresno, CA (7)",   # numbered, weekly-ish
    "Beta Cardiology, Jackson, TN (7 weeks)",          # unnumbered
    "1405    Gamma UK, London UK (7)",                 # EU-shaded below
    "00443    Delta Heart Center, GA (7)",             # no marks -> inactive
    "00460    Epsilon Med., FL (MIX)",                 # cadence from gaps
]
EU_COLUMN_HEADER = "1405    Gamma UK, London UK (7)"

AGGREGATE_HEADERS = [
    "Weeks #", "MFG Date \n(Holidays)", "Calibration date", "Problem?",
    "Month", "FY", "US Demand", "RoW Demand", "Total Commercial", "QC GEN",
    "US\nSTAB",
]


def _workbook(
    marks: dict[str, list[int]] | None = None,
    year: int = 2026,
    weeks: int = 30,
) -> bytes:
    """Build a Master Planner-shaped workbook.

    ``marks`` maps a customer header to the week numbers holding a scheduled
    generator. Defaults give Alpha a 7-week cadence, Beta a 7-week cadence,
    Gamma a 7-week cadence, Delta nothing, and Epsilon a 9-week cadence.
    """
    if marks is None:
        marks = {
            CUSTOMER_HEADERS[0]: [1, 8, 15, 22, 29],
            CUSTOMER_HEADERS[1]: [3, 10, 17, 24],
            CUSTOMER_HEADERS[2]: [2, 9, 16, 23],
            CUSTOMER_HEADERS[3]: [],
            CUSTOMER_HEADERS[4]: [4, 13, 22],
        }

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Schedule"
    ws.append([None] * (len(AGGREGATE_HEADERS) + len(CUSTOMER_HEADERS)))  # spacer
    ws.append(AGGREGATE_HEADERS + CUSTOMER_HEADERS)

    # Shade the EU-restricted header the same way the real workbook does.
    eu_index = len(AGGREGATE_HEADERS) + CUSTOMER_HEADERS.index(EU_COLUMN_HEADER) + 1
    ws.cell(row=2, column=eu_index).fill = PatternFill(
        patternType="solid", start_color=EU_RESTRICTED_FILL,
        end_color=EU_RESTRICTED_FILL,
    )

    monday = datetime(year, 1, 5)
    for week in range(1, weeks + 1):
        mfg = datetime.fromordinal(monday.toordinal() + 7 * (week - 1))
        row = [
            week, mfg, datetime.fromordinal(mfg.toordinal() + 4), None,
            mfg.strftime("%b"), year, 0, 0,
            sum(1 for h in CUSTOMER_HEADERS if week in marks.get(h, [])),
            1, None,
        ]
        row += [1 if week in marks.get(h, []) else None for h in CUSTOMER_HEADERS]
        ws.append(row)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


@pytest.fixture
def workbook_bytes() -> bytes:
    return _workbook()


@pytest.fixture
def derived(workbook_bytes):
    return MasterPlannerConverter().convert(workbook_bytes, horizon_weeks=30)


# ---------------------------------------------------------------------------
# Extraction
# ---------------------------------------------------------------------------

def test_finds_every_customer_column(derived):
    assert len(derived.sites) == len(CUSTOMER_HEADERS)


def test_skips_aggregate_and_marker_columns(derived):
    headers = {s.column_header for s in derived.sites}
    assert "Total Commercial" not in headers
    assert not any("STAB" in h for h in headers)
    assert "US\nSTAB" in derived.ignored_columns


def test_selects_the_planned_year(derived):
    assert derived.year == 2026


def test_reads_first_demand_week_and_cadence(derived):
    alpha = next(s for s in derived.sites if s.site_code == "00449")
    assert alpha.next_demand_week == 1
    assert alpha.interval_weeks == 7
    assert alpha.deliveries == 5
    assert alpha.active is True


def test_cadence_falls_back_to_observed_gaps(derived):
    epsilon = next(s for s in derived.sites if s.site_code == "00460")
    assert epsilon.interval_weeks == 9  # (MIX) header, marks 9 weeks apart


def test_column_without_marks_is_inactive(derived):
    delta = next(s for s in derived.sites if s.site_code == "00443")
    assert delta.active is False


def test_unnumbered_column_gets_a_generated_code(derived):
    beta = next(s for s in derived.sites
                if s.column_header == "Beta Cardiology, Jackson, TN (7 weeks)")
    assert beta.site_code == assign_stable_id(beta.column_header)
    assert beta.code_source == "generated"


def test_eu_shading_is_detected(derived):
    gamma = next(s for s in derived.sites if s.site_code == "1405")
    assert gamma.eu_restricted is True
    assert gamma.country == "uk"
    # Nothing else in the fixture is shaded.
    assert len(derived.eu_sites) == 1


def test_records_the_spreadsheet_column_letter(derived):
    # First customer column sits immediately after the 11 aggregate columns.
    alpha = next(s for s in derived.sites if s.site_code == "00449")
    assert alpha.column_letter == "L"


def test_explicit_year_override_is_honoured():
    data = _workbook(year=2025)
    derived = MasterPlannerConverter().convert(data, horizon_weeks=30, year=2025)
    assert derived.year == 2025
    assert derived.active_sites


def test_year_with_no_rows_yields_inactive_sites_and_a_warning():
    data = _workbook(year=2026)
    derived = MasterPlannerConverter().convert(data, horizon_weeks=30, year=1999)
    assert derived.active_sites == []
    assert any("No schedule rows" in w for w in derived.warnings)


def test_weeks_beyond_the_horizon_are_ignored():
    data = _workbook(marks={CUSTOMER_HEADERS[0]: [1, 8, 25]}, weeks=30)
    derived = MasterPlannerConverter().convert(data, horizon_weeks=10)
    alpha = next(s for s in derived.sites if s.site_code == "00449")
    assert alpha.deliveries == 2


def test_missing_sheet_is_rejected(workbook_bytes):
    with pytest.raises(ValueError, match="not found"):
        MasterPlannerConverter().convert(workbook_bytes, sheet="Nope")


def test_module_level_wrapper(workbook_bytes):
    derived = convert_master_planner(workbook_bytes, horizon_weeks=30)
    assert len(derived.sites) == len(CUSTOMER_HEADERS)


# ---------------------------------------------------------------------------
# Codes agree with the parser — this is what makes the Comparison tab work
# ---------------------------------------------------------------------------

def test_codes_match_the_master_planner_parser(workbook_bytes):
    derived = MasterPlannerConverter().convert(workbook_bytes, horizon_weeks=30)
    parsed = parse_master_planner(workbook_bytes, horizon_weeks=30)

    converted_codes = {s.site_code for s in derived.sites}
    assert set(parsed.customer_schedule) <= converted_codes


def test_duplicate_account_numbers_stay_unique():
    headers = ["00460  Alpha Med., FL (7)", "00460  ??????"]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Schedule"
    ws.append([None] * (len(AGGREGATE_HEADERS) + 2))
    ws.append(AGGREGATE_HEADERS + headers)
    ws.append([1, datetime(2026, 1, 5), datetime(2026, 1, 9), None, "Jan", 2026,
               0, 0, 1, 1, None, 1, None])
    ws.append([8, datetime(2026, 2, 23), datetime(2026, 2, 27), None, "Feb", 2026,
               0, 0, 1, 1, None, 1, None])
    buffer = io.BytesIO()
    wb.save(buffer)

    derived = MasterPlannerConverter().convert(buffer.getvalue(), horizon_weeks=30)
    codes = [s.site_code for s in derived.sites]
    assert codes == ["00460", "00460-2"]
    assert any("more than one column" in w for w in derived.warnings)


# ---------------------------------------------------------------------------
# Output frames and workbook
# ---------------------------------------------------------------------------

def test_sites_frame_has_the_columns_the_optimizer_needs(derived):
    df = sites_frame(derived.sites)
    for column in ("Site_ID", "Active", "Next_Demand_Week", "Interval_Weeks",
                   "Country"):
        assert column in df.columns
    assert set(df["Active"]) <= {"Y", "N"}
    assert set(df["EU_Restricted"]) <= {"Y", "N"}


def test_sites_frame_site_ids_are_unique(derived):
    df = sites_frame(derived.sites)
    assert df["Site_ID"].is_unique


def test_sites_frame_keeps_leading_zeros(derived):
    df = sites_frame(derived.sites)
    assert "00449" in set(df["Site_ID"])


def test_mapping_frame_links_codes_to_columns(derived):
    df = mapping_frame(derived.sites)
    assert len(df) == len(derived.sites)
    row = df[df["Site_ID"] == "00449"].iloc[0]
    assert row["Master_Planner_Header"].startswith("00449")
    assert row["Master_Planner_Column"] == "L"
    assert row["Code_Source"] == "account number"


def test_notes_frame_lists_one_row_per_note(derived):
    df = notes_frame(derived.sites)
    assert len(df) == sum(len(s.notes) for s in derived.sites)
    assert set(df.columns) == {
        "Site_ID", "Site_Name", "Master_Planner_Column", "Note",
    }


def test_written_workbook_has_the_three_sheets(derived):
    data = MasterPlannerConverter().write(derived)
    wb = openpyxl.load_workbook(io.BytesIO(data))
    try:
        assert wb.sheetnames == ["Sites", "Site_Mapping", "Conversion_Notes"]
    finally:
        wb.close()


def test_written_workbook_reads_back_through_the_sites_reader(derived):
    from io_adapters.sites_reader import ExcelSitesReader

    data = MasterPlannerConverter().write(derived)
    df = ExcelSitesReader().read(data, "Sites")
    assert "site_id" in df.columns
    assert len(df) == len(derived.sites)
    # Leading zeros must survive the round-trip or per-customer matching breaks.
    assert "00449" in set(df["site_id"])
