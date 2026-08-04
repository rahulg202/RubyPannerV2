"""Tests for the unified workbook export (io_adapters.workbook_exporter)."""

import io
from datetime import date

import openpyxl
import pandas as pd
import pytest

from domain.delivery_assignment import DeliveryRecord
from domain.params import IntegratedParams, SupplierParams
from domain.quota import QuarterlyQuotaStatus
from io_adapters.master_planner_parser import AssignedId
from io_adapters.workbook_exporter import (
    FILL_EARLY,
    FILL_LATE,
    FILL_NEW,
    SHEET_ASSIGNED_IDS,
    SHEET_CHANGED,
    SHEET_COMPARISON,
    SHEET_ISSUES,
    SHEET_PARAMS,
    SHEET_PLAN,
    SHEET_QUOTA,
    SHEET_SITES,
    SHEET_WEEKLY_CMP,
    ExportBundle,
    WorkbookExporter,
    write_unified_workbook,
)

PARAMS = IntegratedParams()
SUPPLIER = SupplierParams()


def _plan():
    return pd.DataFrame({
        "Week": [1, 2, 3],
        "Good_Production": [15, 30, 0],
        "Net_Inventory_End": [0, 0, 0],
    })


def _sites():
    return pd.DataFrame({"site_id": ["S1"], "next_demand_week": [1], "interval_weeks": [7]})


def _issues():
    return pd.DataFrame(columns=["row_index", "site_id", "issue"])


def _summary():
    return {
        "total_composite_cost": 1000.0, "total_penalty_cost": 500.0,
        "total_overtime_cost": 300.0, "total_capacity_cost": 200.0,
        "overtime_weeks": 1, "total_quota_penalty_cost": 50.0,
        "w_penalty": 1.0, "w_overtime": 1.0, "w_capacity": 1.0,
    }


def _assignments():
    return [
        DeliveryRecord("S1", "Acme", "usa", due_week=3, planned_week=1,
                       due_week_shift=-2, manual_week=3, week_shift=-2,
                       is_early=True, compared=True),
        DeliveryRecord("S2", "Beta", "usa", due_week=1, planned_week=2,
                       due_week_shift=1, manual_week=1, week_shift=1,
                       is_late=True, compared=True),
        DeliveryRecord("S3", "Gamma", "usa", due_week=2, planned_week=2,
                       due_week_shift=0, manual_week=2, week_shift=0,
                       compared=True),
        DeliveryRecord("S4", "Delta", "uk", due_week=2, planned_week=2,
                       due_week_shift=0, is_new_customer=True, compared=True),
    ]


def _quota(shortfall=True):
    return [
        QuarterlyQuotaStatus(
            "Curium", 1, (1, 2, 3), 10000.0, 12000.0, -2000.0, 0.0, 0.0,
            weeks_covered=13, expected_weeks=13, status="OK",
        ),
        QuarterlyQuotaStatus(
            "BWXT", 1, (1, 2, 3), 10000.0, 4000.0, 6000.0,
            6000.0 if shortfall else 0.0, 300.0 if shortfall else 0.0,
            weeks_covered=13, expected_weeks=13,
            status="SHORTFALL" if shortfall else "OK",
        ),
    ]


def _partial_quota():
    """A partial quarter: reported, pro-rated target, no penalty."""
    return [
        QuarterlyQuotaStatus(
            "Curium", 2, (14, 15, 16), 10000.0, 1200.0, 8800.0, 0.0, 0.0,
            is_partial=True, weeks_covered=3, expected_weeks=13,
            prorated_quota_mci=10000.0 * 3 / 13,
            prorated_shortfall_mci=10000.0 * 3 / 13 - 1200.0,
            status="Partial — not penalised",
        ),
    ]


def _bundle(**over):
    base = dict(
        plan_df=_plan(), sites_df=_sites(), issues_df=_issues(),
        params=PARAMS, summary=_summary(),
    )
    base.update(over)
    return ExportBundle(**base)


def _sheets(data: bytes):
    return pd.ExcelFile(io.BytesIO(data)).sheet_names


# ---------------------------------------------------------------------------
# Core sheets
# ---------------------------------------------------------------------------

def test_core_sheets_always_present():
    data = write_unified_workbook(_bundle())
    assert {SHEET_PLAN, SHEET_SITES, SHEET_ISSUES, SHEET_PARAMS} <= set(_sheets(data))


def test_optional_sheets_absent_when_no_data():
    data = write_unified_workbook(_bundle())
    names = set(_sheets(data))
    assert SHEET_CHANGED not in names
    assert SHEET_QUOTA not in names
    assert SHEET_COMPARISON not in names
    assert SHEET_ASSIGNED_IDS not in names


def test_all_sheets_present_when_fully_populated():
    data = write_unified_workbook(_bundle(
        supplier_params=SUPPLIER,
        quota_status=_quota(),
        assignments=_assignments(),
        week_dates=[(1, date(2026, 1, 5), date(2026, 1, 9)),
                    (2, date(2026, 1, 12), date(2026, 1, 16)),
                    (3, date(2026, 1, 19), date(2026, 1, 23))],
        comparison_components=[{
            "Component": "Penalty", "Baseline": 100.0, "Optimized": 40.0,
            "Saving_Abs": 60.0, "Saving_Pct": 60.0,
        }],
        weekly_comparison=pd.DataFrame({
            "Week": [1], "Manual_Production": [10],
            "Optimized_Production": [12], "Difference": [2],
        }),
        assigned_ids=[AssignedId("RF-abc12345", "Apex Cardiology", "Apex Cardiology (7)", "apex cardiology 7")],
    ))
    assert set(_sheets(data)) == {
        SHEET_PLAN, SHEET_SITES, SHEET_ISSUES, SHEET_PARAMS,
        SHEET_CHANGED, SHEET_QUOTA, SHEET_COMPARISON,
        SHEET_WEEKLY_CMP, SHEET_ASSIGNED_IDS,
    }


# ---------------------------------------------------------------------------
# Date columns
# ---------------------------------------------------------------------------

def test_date_columns_inserted_after_week():
    data = write_unified_workbook(_bundle(week_dates=[
        (1, date(2026, 1, 5), date(2026, 1, 9)),
        (2, date(2026, 1, 12), date(2026, 1, 16)),
        (3, date(2026, 1, 19), date(2026, 1, 23)),
    ]))
    df = pd.read_excel(io.BytesIO(data), sheet_name=SHEET_PLAN)
    assert list(df.columns)[:3] == ["Week", "MFG_Date", "Cal_Date"]
    assert pd.Timestamp(df.iloc[0]["MFG_Date"]).date() == date(2026, 1, 5)
    assert pd.Timestamp(df.iloc[0]["Cal_Date"]).date() == date(2026, 1, 9)


def test_no_date_columns_without_week_dates():
    data = write_unified_workbook(_bundle())
    df = pd.read_excel(io.BytesIO(data), sheet_name=SHEET_PLAN)
    assert "MFG_Date" not in df.columns


# ---------------------------------------------------------------------------
# Model_Params
# ---------------------------------------------------------------------------

def test_supplier_params_recorded():
    data = write_unified_workbook(_bundle(supplier_params=SUPPLIER))
    mp = pd.read_excel(io.BytesIO(data), sheet_name=SHEET_PARAMS)
    names = set(mp["Parameter"])
    assert {"curium_surplus_pct", "bwxt_surplus_pct", "first_run_allocation",
            "curium_quarterly_quota_mci", "quota_shortfall_penalty_rate",
            "quarter_start_month"} <= names


def test_reference_week_and_offset_recorded():
    data = write_unified_workbook(_bundle(
        reference_week_date=date(2026, 1, 5), calibration_offset_days=4
    ))
    mp = pd.read_excel(io.BytesIO(data), sheet_name=SHEET_PARAMS)
    names = set(mp["Parameter"])
    assert {"reference_week_date", "calibration_offset_days"} <= names


def test_quota_penalty_in_params_summary():
    data = write_unified_workbook(_bundle())
    mp = pd.read_excel(io.BytesIO(data), sheet_name=SHEET_PARAMS)
    assert "total_quota_penalty_cost_usd" in set(mp["Parameter"])


def test_row_cap_described_as_qc_cap():
    data = write_unified_workbook(_bundle())
    mp = pd.read_excel(io.BytesIO(data), sheet_name=SHEET_PARAMS)
    desc = mp[mp["Parameter"] == "row_cap"].iloc[0]["Description"]
    assert "QC" in desc


# ---------------------------------------------------------------------------
# Changed weeks + highlighting
# ---------------------------------------------------------------------------

def test_changed_weeks_content():
    data = write_unified_workbook(_bundle(assignments=_assignments()))
    df = pd.read_excel(io.BytesIO(data), sheet_name=SHEET_CHANGED)
    assert len(df) == 4
    assert set(["Site_ID", "Manual_Plan_Week", "Optimized_Week", "Week_Shift",
                "Due_Week", "Is_New_Customer"]) <= set(df.columns)
    s1 = df[df["Site_ID"] == "S1"].iloc[0]
    assert s1["Week_Shift"] == -2
    assert s1["Manual_Plan_Week"] == 3


def test_changed_week_highlighting_colours():
    data = write_unified_workbook(_bundle(assignments=_assignments()))
    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb[SHEET_CHANGED]
    # Row order matches the assignments list: early, late, on-time, new
    early = ws.cell(row=2, column=1).fill.start_color.rgb
    late = ws.cell(row=3, column=1).fill.start_color.rgb
    ontime = ws.cell(row=4, column=1).fill.start_color.rgb
    new = ws.cell(row=5, column=1).fill.start_color.rgb
    assert early == FILL_EARLY.start_color.rgb
    assert late == FILL_LATE.start_color.rgb
    assert new == FILL_NEW.start_color.rgb
    assert ontime not in (FILL_EARLY.start_color.rgb, FILL_LATE.start_color.rgb,
                          FILL_NEW.start_color.rgb)


def test_changed_weeks_include_dates_when_available():
    data = write_unified_workbook(_bundle(
        assignments=_assignments(),
        week_dates=[(1, date(2026, 1, 5), date(2026, 1, 9)),
                    (2, date(2026, 1, 12), date(2026, 1, 16)),
                    (3, date(2026, 1, 19), date(2026, 1, 23))],
    ))
    df = pd.read_excel(io.BytesIO(data), sheet_name=SHEET_CHANGED)
    assert "Optimized_MFG_Date" in df.columns


# ---------------------------------------------------------------------------
# Quota
# ---------------------------------------------------------------------------

def test_quota_sheet_content():
    data = write_unified_workbook(_bundle(quota_status=_quota()))
    df = pd.read_excel(io.BytesIO(data), sheet_name=SHEET_QUOTA)
    assert set(df["Supplier"]) == {"Curium", "BWXT"}
    bwxt = df[df["Supplier"] == "BWXT"].iloc[0]
    assert bwxt["Gap_mCi"] == 6000.0
    assert bwxt["Penalty_USD"] == 300.0
    assert bwxt["Status"] == "SHORTFALL"


def test_quota_sheet_has_coverage_columns():
    data = write_unified_workbook(_bundle(quota_status=_quota()))
    df = pd.read_excel(io.BytesIO(data), sheet_name=SHEET_QUOTA)
    for column in ("Weeks_Covered", "Weeks_In_Quarter", "Is_Partial",
                   "Quota_mCi", "Target_mCi", "Gap_mCi", "Status"):
        assert column in df.columns


def test_quota_shortfall_row_highlighted():
    data = write_unified_workbook(_bundle(quota_status=_quota()))
    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb[SHEET_QUOTA]
    # Row 2 = Curium (no shortfall), row 3 = BWXT (penalised shortfall)
    assert ws.cell(row=3, column=1).fill.start_color.rgb == FILL_LATE.start_color.rgb
    assert ws.cell(row=2, column=1).fill.start_color.rgb != FILL_LATE.start_color.rgb


def test_partial_quarter_row_not_highlighted_and_prorated():
    """A partial quarter has a gap but no penalty, so it must not look like a breach."""
    data = write_unified_workbook(_bundle(quota_status=_partial_quota()))
    df = pd.read_excel(io.BytesIO(data), sheet_name=SHEET_QUOTA)
    row = df.iloc[0]
    assert row["Is_Partial"] == "Y"
    assert row["Weeks_Covered"] == 3
    assert row["Weeks_In_Quarter"] == 13
    assert row["Penalty_USD"] == 0.0
    assert row["Target_mCi"] < row["Quota_mCi"]      # pro-rated target

    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb[SHEET_QUOTA]
    assert ws.cell(row=2, column=1).fill.start_color.rgb != FILL_LATE.start_color.rgb


# ---------------------------------------------------------------------------
# Comparison and assigned IDs
# ---------------------------------------------------------------------------

def test_comparison_sheet_handles_none_pct():
    data = write_unified_workbook(_bundle(comparison_components=[
        {"Component": "Penalty", "Baseline": 0.0, "Optimized": 0.0,
         "Saving_Abs": 0.0, "Saving_Pct": None},
    ]))
    # keep_default_na=False: pandas otherwise treats the literal "n/a" as a
    # missing-value token on read. The cell genuinely contains the text "n/a",
    # which is what a planner sees in Excel.
    df = pd.read_excel(
        io.BytesIO(data), sheet_name=SHEET_COMPARISON, keep_default_na=False
    )
    assert df.iloc[0]["Saving_Pct"] == "n/a"


def test_comparison_sheet_formats_pct_as_text():
    data = write_unified_workbook(_bundle(comparison_components=[
        {"Component": "Penalty", "Baseline": 100.0, "Optimized": 40.0,
         "Saving_Abs": 60.0, "Saving_Pct": 60.0},
        {"Component": "Overtime", "Baseline": 0.0, "Optimized": 0.0,
         "Saving_Abs": 0.0, "Saving_Pct": None},
    ]))
    df = pd.read_excel(
        io.BytesIO(data), sheet_name=SHEET_COMPARISON, keep_default_na=False
    )
    assert df.iloc[0]["Saving_Pct"] == "60.00%"
    assert df.iloc[1]["Saving_Pct"] == "n/a"


def test_comparison_pct_cell_written_as_text_in_workbook():
    """Verify at the cell level, independent of pandas NA handling."""
    data = write_unified_workbook(_bundle(comparison_components=[
        {"Component": "Penalty", "Baseline": 0.0, "Optimized": 0.0,
         "Saving_Abs": 0.0, "Saving_Pct": None},
    ]))
    ws = openpyxl.load_workbook(io.BytesIO(data))[SHEET_COMPARISON]
    header = [c.value for c in ws[1]]
    col = header.index("Saving_Pct") + 1
    assert ws.cell(row=2, column=col).value == "n/a"


def test_assigned_ids_sheet_content():
    data = write_unified_workbook(_bundle(assigned_ids=[
        AssignedId("RF-abc12345", "Apex Cardiology", "Apex Cardiology (7)", "apex cardiology 7"),
    ]))
    df = pd.read_excel(io.BytesIO(data), sheet_name=SHEET_ASSIGNED_IDS)
    assert list(df.columns) == ["Generated_ID", "Customer_Name", "Master_Planner_Header"]
    assert df.iloc[0]["Generated_ID"] == "RF-abc12345"


# ---------------------------------------------------------------------------
# Exporter class / port compatibility
# ---------------------------------------------------------------------------

def test_exporter_class_accepts_extras():
    data = WorkbookExporter().export(
        _plan(), _sites(), _issues(), PARAMS, _summary(),
        supplier_params=SUPPLIER, quota_status=_quota(), assignments=_assignments(),
    )
    names = set(_sheets(data))
    assert SHEET_QUOTA in names and SHEET_CHANGED in names


def test_exporter_class_without_extras_is_core_only():
    data = WorkbookExporter().export(_plan(), _sites(), _issues(), PARAMS, _summary())
    assert set(_sheets(data)) == {SHEET_PLAN, SHEET_SITES, SHEET_ISSUES, SHEET_PARAMS}


def test_exporter_still_satisfies_port():
    from services.ports import ResultExporterPort
    assert isinstance(WorkbookExporter(), ResultExporterPort)
