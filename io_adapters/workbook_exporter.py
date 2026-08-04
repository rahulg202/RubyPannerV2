"""Adapter: write optimization results to an Excel workbook.

Canonical home for result export. Produces the unified workbook covering every
output of a run: the weekly plan (with calendar dates and supplier allocation),
model parameters, changed customer weeks (with colour highlighting), the
manual-vs-optimized cost comparison, quarterly supplier quota status, and the
generated-identifier mapping.

The legacy four-sheet ``export_excel`` is retained for the CLI. ``WorkbookExporter``
implements :class:`services.ports.ResultExporterPort` and returns bytes for the
Streamlit download flow.
"""

from __future__ import annotations

import io
from dataclasses import dataclass, field
from typing import Any, Sequence

import pandas as pd
from openpyxl.styles import PatternFill

from domain.params import IntegratedParams, SupplierParams

# Sheet names
SHEET_PLAN = "Weekly_Plan"
SHEET_SITES = "Sites_Clean"
SHEET_ISSUES = "Input_Issues"
SHEET_PARAMS = "Model_Params"
SHEET_CHANGED = "Changed_Weeks"
SHEET_COMPARISON = "Cost_Comparison"
SHEET_WEEKLY_CMP = "Weekly_Comparison"
SHEET_QUOTA = "Quota_Status"
SHEET_ASSIGNED_IDS = "Assigned_IDs"

# Highlight fills. Chosen to avoid the Master Planner's existing dark blue
# (FF002060), yellow (FFFF00), orange (FFC000), and red (FF0000) conventions.
FILL_EARLY = PatternFill(start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid")
FILL_LATE = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")
FILL_NEW = PatternFill(start_color="FFBDD7EE", end_color="FFBDD7EE", fill_type="solid")
FILL_SHORTFALL = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")


# ---------------------------------------------------------------------------
# Model_Params rows
# ---------------------------------------------------------------------------

def _build_params_rows(
    params: IntegratedParams,
    summary: dict,
    supplier_params: SupplierParams | None = None,
    reference_week_date: Any = None,
    calibration_offset_days: int | None = None,
) -> list[tuple]:
    """Build the (Parameter, Value, Description) rows for the Model_Params sheet."""
    rows = [
        ("horizon_weeks", params.horizon_weeks, "Planning horizon in weeks"),
        ("penalty_rate", params.penalty_rate, "USD per unit-week early inventory"),
        ("late_penalty_multiplier", params.late_penalty_multiplier, "Multiplier on penalty_rate for backlog"),
        ("late_penalty_rate", params.late_penalty_rate, "Effective backlog penalty rate (derived)"),
        ("overtime_rate", params.overtime_rate, "USD per overtime week (3rd batch)"),
        ("capacity_rate", params.capacity_rate, "USD per unused good unit slot per week"),
        ("w_penalty", params.w_penalty, "Weight for penalty cost component"),
        ("w_overtime", params.w_overtime, "Weight for overtime cost component"),
        ("w_capacity", params.w_capacity, "Weight for capacity utilization cost component"),
        ("row_cap", params.row_cap, "QC shipping cap: max restricted-country units per week"),
        ("min_batch_produced", params.min_batch_produced, "Min units produced per batch (incl. test discard)"),
        ("max_batch_produced", params.max_batch_produced, "Max units produced per batch (incl. test discard)"),
        ("test_discard_per_batch", params.test_discard_per_batch, "Test units discarded per batch"),
        ("normal_max_batches", params.normal_max_batches, "Max batches in a normal week"),
        ("overtime_max_batches", params.overtime_max_batches, "Max batches in an overtime week"),
        ("normal_max_good_week", params.normal_max_good_week, "Max good units in a normal week (derived)"),
        ("overtime_max_good_week", params.overtime_max_good_week, "Max good units in an overtime week (derived)"),
    ]

    if reference_week_date is not None:
        rows.append(("reference_week_date", str(reference_week_date),
                     "Calendar manufacturing date anchoring planning week 1"))
    if calibration_offset_days is not None:
        rows.append(("calibration_offset_days", calibration_offset_days,
                     "Days from manufacturing date to calibration date"))

    if supplier_params is not None:
        sp = supplier_params
        rows.extend([
            ("per_generator_mci", sp.per_generator_mci, "Sr-82 mCi per good generator"),
            ("per_batch_mci", sp.per_batch_mci, "Sr-82 mCi per batch (one QC generator)"),
            ("minimum_surplus_mci", sp.minimum_surplus_mci, "Floor on the surplus term (mCi)"),
            ("curium_surplus_pct", sp.curium_surplus_pct, "Curium surplus fraction"),
            ("bwxt_surplus_pct", sp.bwxt_surplus_pct, "BWXT surplus fraction"),
            ("first_run_allocation", sp.first_run_allocation, "Generators in the first Curium run of a split week"),
            ("curium_quarterly_quota_mci", sp.curium_quarterly_quota_mci, "Curium minimum quota per quarter (mCi)"),
            ("bwxt_quarterly_quota_mci", sp.bwxt_quarterly_quota_mci, "BWXT minimum quota per quarter (mCi)"),
            ("quota_shortfall_penalty_rate", sp.quota_shortfall_penalty_rate, "USD per mCi of quarterly quota shortfall"),
            ("w_quota", sp.w_quota, "Weight for quota shortfall cost component"),
            ("quarter_start_month", sp.quarter_start_month, "Month on which quarter 1 begins"),
            ("curium_unavailable_weeks", ", ".join(map(str, sp.curium_unavailable_weeks)) or "(none)",
             "Weeks Curium cannot supply"),
            ("bwxt_unavailable_weeks", ", ".join(map(str, sp.bwxt_unavailable_weeks)) or "(none)",
             "Weeks BWXT cannot supply"),
        ])

    rows.extend([
        ("total_composite_cost_usd", summary.get("total_composite_cost", ""), "Total composite cost across horizon"),
        ("total_penalty_cost_usd", summary.get("total_penalty_cost", ""), "Total penalty cost across horizon"),
        ("total_overtime_cost_usd", summary.get("total_overtime_cost", ""), "Total overtime cost across horizon"),
        ("total_capacity_cost_usd", summary.get("total_capacity_cost", ""), "Total capacity utilization cost across horizon"),
        ("overtime_weeks", summary.get("overtime_weeks", ""), "Number of weeks with 3rd batch"),
    ])
    if "total_quota_penalty_cost" in summary:
        rows.append(("total_quota_penalty_cost_usd", summary["total_quota_penalty_cost"],
                     "Total quarterly quota shortfall penalty"))
    return rows


# ---------------------------------------------------------------------------
# Legacy four-sheet export (CLI)
# ---------------------------------------------------------------------------

def _write_workbook(target, plan_df, active_df, issues_df, params, summary) -> None:
    """Write the four standard sheets to ``target`` (a path or file-like buffer)."""
    params_df = pd.DataFrame(
        _build_params_rows(params, summary),
        columns=["Parameter", "Value", "Description"],
    )
    with pd.ExcelWriter(target, engine="openpyxl") as writer:
        plan_df.to_excel(writer, sheet_name=SHEET_PLAN, index=False)
        active_df.to_excel(writer, sheet_name=SHEET_SITES, index=False)
        issues_df.to_excel(writer, sheet_name=SHEET_ISSUES, index=False)
        params_df.to_excel(writer, sheet_name=SHEET_PARAMS, index=False)


def export_excel(
    output_path: str,
    plan_df: pd.DataFrame,
    active_df: pd.DataFrame,
    issues_df: pd.DataFrame,
    params: IntegratedParams,
    summary: dict,
) -> None:
    """Write the four core sheets to an Excel file (used by the CLI)."""
    _write_workbook(output_path, plan_df, active_df, issues_df, params, summary)


# ---------------------------------------------------------------------------
# Unified export
# ---------------------------------------------------------------------------

@dataclass
class ExportBundle:
    """Everything the unified workbook can contain. Optional parts may be None."""

    plan_df: pd.DataFrame
    sites_df: pd.DataFrame
    issues_df: pd.DataFrame
    params: IntegratedParams
    summary: dict
    supplier_params: SupplierParams | None = None
    quota_status: Sequence[Any] = field(default_factory=tuple)
    assignments: Sequence[Any] = field(default_factory=tuple)
    week_dates: Sequence[tuple] = field(default_factory=tuple)
    comparison_components: Sequence[dict] = field(default_factory=tuple)
    weekly_comparison: pd.DataFrame | None = None
    assigned_ids: Sequence[Any] = field(default_factory=tuple)
    reference_week_date: Any = None
    calibration_offset_days: int | None = None


def _with_dates(plan_df: pd.DataFrame, week_dates: Sequence[tuple]) -> pd.DataFrame:
    """Insert MFG_Date and Cal_Date immediately after Week, when available."""
    if not week_dates or "Week" not in plan_df.columns:
        return plan_df
    mfg = {w: m for w, m, _c in week_dates}
    cal = {w: c for w, _m, c in week_dates}
    out = plan_df.copy()
    out.insert(1, "MFG_Date", out["Week"].map(mfg))
    out.insert(2, "Cal_Date", out["Week"].map(cal))
    return out


def _quota_frame(quota_status: Sequence[Any]) -> pd.DataFrame:
    """Build the Quota_Status sheet.

    ``Target_mCi`` is the full quota for a complete quarter, or the pro-rated
    run-rate target for a partial one. Partial quarters carry no penalty.
    """
    return pd.DataFrame([
        {
            "Supplier": q.supplier,
            "Quarter": q.quarter,
            "Weeks": f"{min(q.weeks)}-{max(q.weeks)}" if q.weeks else "",
            "Weeks_Covered": q.weeks_covered,
            "Weeks_In_Quarter": q.expected_weeks,
            "Is_Partial": "Y" if q.is_partial else "N",
            "Quota_mCi": q.quota_mci,
            "Target_mCi": q.prorated_quota_mci if q.is_partial else q.quota_mci,
            "Ordered_mCi": q.ordered_mci,
            "Gap_mCi": (q.prorated_shortfall_mci if q.is_partial
                        else q.shortfall_mci),
            "Penalty_USD": q.penalty_usd,
            "Status": q.status,
        }
        for q in quota_status
    ])


def _changed_frame(assignments: Sequence[Any], week_dates: Sequence[tuple]) -> pd.DataFrame:
    mfg = {w: m for w, m, _c in week_dates} if week_dates else {}
    cal = {w: c for w, _m, c in week_dates} if week_dates else {}
    rows = []
    for r in assignments:
        row = {
            "Site_ID": r.site_id,
            "Site_Name": r.site_name,
            "Country": r.country,
            "Manual_Plan_Week": r.manual_week,
            "Optimized_Week": r.planned_week,
            "Week_Shift": r.week_shift,
            "Due_Week": r.due_week,
            "Is_New_Customer": "Y" if r.is_new_customer else "N",
        }
        if mfg:
            row["Optimized_MFG_Date"] = mfg.get(r.planned_week)
            row["Optimized_Cal_Date"] = cal.get(r.planned_week)
        rows.append(row)
    return pd.DataFrame(rows)


def _assigned_ids_frame(assigned_ids: Sequence[Any]) -> pd.DataFrame:
    return pd.DataFrame([
        {
            "Generated_ID": a.generated_id,
            "Customer_Name": a.customer_name,
            "Master_Planner_Header": a.column_header,
        }
        for a in assigned_ids
    ])


def _comparison_frame(components: Sequence[dict]) -> pd.DataFrame:
    """Build the Cost_Comparison sheet.

    ``Saving_Pct`` is written as text so an undefined percentage (zero baseline)
    shows as an explicit "n/a" rather than being coerced to a blank NaN cell.
    """
    df = pd.DataFrame([
        {
            "Component": c["Component"],
            "Baseline": c["Baseline"],
            "Optimized": c["Optimized"],
            "Saving_Abs": c["Saving_Abs"],
            "Saving_Pct": (
                "n/a" if c["Saving_Pct"] is None else f"{c['Saving_Pct']:.2f}%"
            ),
        }
        for c in components
    ])
    if not df.empty:
        df["Saving_Pct"] = df["Saving_Pct"].astype(str)
    return df


def _highlight_changed_weeks(worksheet, changed_df: pd.DataFrame) -> None:
    """Fill rows: green moved earlier, red moved later, blue newly-added."""
    if changed_df.empty:
        return
    cols = list(changed_df.columns)
    shift_idx = cols.index("Week_Shift")
    new_idx = cols.index("Is_New_Customer")
    width = len(cols)

    for offset, (_i, row) in enumerate(changed_df.iterrows()):
        excel_row = offset + 2  # header occupies row 1
        shift = row.iloc[shift_idx]
        is_new = str(row.iloc[new_idx]).upper() == "Y"
        if is_new:
            fill = FILL_NEW
        elif shift is None or pd.isna(shift):
            continue  # no manual counterpart to compare against
        elif shift < 0:
            fill = FILL_EARLY
        elif shift > 0:
            fill = FILL_LATE
        else:
            continue  # same week as the manual plan: no fill
        for col in range(1, width + 1):
            worksheet.cell(row=excel_row, column=col).fill = fill


def _highlight_quota_shortfall(worksheet, quota_df: pd.DataFrame) -> None:
    """Fill rows with a genuine, penalised shortfall.

    Partial quarters are never highlighted: their gap is a run-rate reference,
    not a compliance failure.
    """
    if quota_df.empty or "Penalty_USD" not in quota_df.columns:
        return
    cols = list(quota_df.columns)
    penalty_idx = cols.index("Penalty_USD")
    width = len(cols)
    for offset, (_i, row) in enumerate(quota_df.iterrows()):
        if float(row.iloc[penalty_idx]) <= 0:
            continue
        excel_row = offset + 2
        for col in range(1, width + 1):
            worksheet.cell(row=excel_row, column=col).fill = FILL_SHORTFALL


def write_unified_workbook(bundle: ExportBundle) -> bytes:
    """Write every available section into one workbook and return the bytes."""
    plan_df = _with_dates(bundle.plan_df, bundle.week_dates)
    params_df = pd.DataFrame(
        _build_params_rows(
            bundle.params, bundle.summary, bundle.supplier_params,
            bundle.reference_week_date, bundle.calibration_offset_days,
        ),
        columns=["Parameter", "Value", "Description"],
    )
    changed_df = _changed_frame(bundle.assignments, bundle.week_dates)
    quota_df = _quota_frame(bundle.quota_status)

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        plan_df.to_excel(writer, sheet_name=SHEET_PLAN, index=False)
        bundle.sites_df.to_excel(writer, sheet_name=SHEET_SITES, index=False)
        bundle.issues_df.to_excel(writer, sheet_name=SHEET_ISSUES, index=False)
        params_df.to_excel(writer, sheet_name=SHEET_PARAMS, index=False)

        if not changed_df.empty:
            changed_df.to_excel(writer, sheet_name=SHEET_CHANGED, index=False)
            _highlight_changed_weeks(writer.sheets[SHEET_CHANGED], changed_df)

        if not quota_df.empty:
            quota_df.to_excel(writer, sheet_name=SHEET_QUOTA, index=False)
            _highlight_quota_shortfall(writer.sheets[SHEET_QUOTA], quota_df)

        if bundle.comparison_components:
            _comparison_frame(bundle.comparison_components).to_excel(
                writer, sheet_name=SHEET_COMPARISON, index=False
            )

        if bundle.weekly_comparison is not None and not bundle.weekly_comparison.empty:
            bundle.weekly_comparison.to_excel(
                writer, sheet_name=SHEET_WEEKLY_CMP, index=False
            )

        if bundle.assigned_ids:
            _assigned_ids_frame(bundle.assigned_ids).to_excel(
                writer, sheet_name=SHEET_ASSIGNED_IDS, index=False
            )

    return buffer.getvalue()


class WorkbookExporter:
    """Writes the result workbook to bytes. Implements ``ResultExporterPort``."""

    def export(
        self,
        plan_df: pd.DataFrame,
        sites_df: pd.DataFrame,
        issues_df: pd.DataFrame,
        params: IntegratedParams,
        summary: dict,
        **extras: Any,
    ) -> bytes:
        """Export a workbook. Extra sections are included when supplied.

        Accepted ``extras``: ``supplier_params``, ``quota_status``,
        ``assignments``, ``week_dates``, ``comparison_components``,
        ``weekly_comparison``, ``assigned_ids``, ``reference_week_date``,
        ``calibration_offset_days``.
        """
        bundle = ExportBundle(
            plan_df=plan_df,
            sites_df=sites_df,
            issues_df=issues_df,
            params=params,
            summary=summary,
            supplier_params=extras.get("supplier_params"),
            quota_status=extras.get("quota_status") or (),
            assignments=extras.get("assignments") or (),
            week_dates=extras.get("week_dates") or (),
            comparison_components=extras.get("comparison_components") or (),
            weekly_comparison=extras.get("weekly_comparison"),
            assigned_ids=extras.get("assigned_ids") or (),
            reference_week_date=extras.get("reference_week_date"),
            calibration_offset_days=extras.get("calibration_offset_days"),
        )
        return write_unified_workbook(bundle)
