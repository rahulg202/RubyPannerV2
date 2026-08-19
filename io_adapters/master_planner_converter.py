"""Adapter: turn the wide Master Planner into an optimizer-ready input file.

The planning team's manual plan lives in one wide sheet — a row per production
week, a column per customer, ``1`` in a cell for a scheduled generator. Building
the optimizer's input sheet from it by hand was slow and error-prone, and the two
files shared no key, so the Comparison tool could not match sites between them.

This adapter reads the Master Planner and writes a workbook with three sheets:

``Sites``
    Ready to upload to the Cost Optimizer as-is.
``Site_Mapping``
    Every site code next to the Master Planner column it came from, so the codes
    can be pasted back into the headers and both files stay linked.
``Conversion_Notes``
    Anything the planner should check — missing account numbers, cadences that
    disagree with the schedule, columns with no scheduled work.

Site codes reuse the Master Planner parser's scheme (leading account number, else
``RF-<hash8>``) so the Comparison tool matches without further work. Only this
layer touches openpyxl; the derivation rules are pure and live in
``domain.site_derivation``.
"""

from __future__ import annotations

import io
from dataclasses import dataclass, field

import openpyxl
import pandas as pd
from openpyxl.utils import get_column_letter

from domain.site_derivation import (
    DerivedSite,
    PlannerColumn,
    derive_sites,
    mapping_frame,
    notes_frame,
    sites_frame,
)
from io_adapters.master_planner_parser import (
    HEADER_ROW,
    MasterPlannerParser,
    _is_customer_column,
    _is_schedule_mark,
    _to_int,
    assign_stable_id,
)

# Fill used in the Master Planner to shade EU-restricted customers
# (Denmark, the Netherlands and the UK — the sites that must ship from Curium).
EU_RESTRICTED_FILL = "FF002060"

SITES_SHEET = "Sites"
MAPPING_SHEET = "Site_Mapping"
NOTES_SHEET = "Conversion_Notes"


@dataclass
class DerivedSiteSet:
    """Every site derived from one Master Planner workbook, plus provenance."""

    sites: list[DerivedSite] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    year: int | None = None
    horizon_weeks: int = 52
    ignored_columns: list[str] = field(default_factory=list)

    @property
    def active_sites(self) -> list[DerivedSite]:
        return [s for s in self.sites if s.active]

    @property
    def generated_codes(self) -> list[DerivedSite]:
        return [s for s in self.sites if s.code_source == "generated"]

    @property
    def eu_sites(self) -> list[DerivedSite]:
        return [s for s in self.sites if s.eu_restricted]

    @property
    def flagged_sites(self) -> list[DerivedSite]:
        return [s for s in self.sites if s.notes]


def _fill_rgb(cell) -> str | None:
    """Return the cell's solid fill colour as an uppercase RGB string, if any."""
    fill = cell.fill
    if fill is None or fill.patternType != "solid":
        return None
    rgb = getattr(fill.start_color, "rgb", None)
    if rgb is None:
        return None
    text = str(rgb)
    # Theme-coloured cells yield a non-RGB placeholder rather than a hex string.
    return text.upper() if len(text) in (6, 8) else None


class MasterPlannerConverter:
    """Reads the Master Planner and produces optimizer-ready site rows.

    Implements :class:`services.ports.MasterPlannerConverterPort`.
    """

    def convert(
        self,
        source: bytes,
        sheet: str = "Schedule",
        horizon_weeks: int = 52,
        year: int | None = None,
    ) -> DerivedSiteSet:
        """Derive one site per customer column.

        Parameters
        ----------
        source : bytes
            Raw Master Planner workbook bytes.
        sheet : str
            Worksheet holding the schedule grid (default ``Schedule``).
        horizon_weeks : int
            Planning horizon. Weeks outside ``1..horizon_weeks`` are ignored.
        year : int | None
            Which calendar year of rows to read. Week numbers repeat per year in
            this workbook, so a window is required; ``None`` picks the most
            completely-planned year, matching the Master Planner parser.

        Raises
        ------
        ValueError
            If the sheet is missing or has no recognisable header row.
        """
        # Styles are needed for the EU-restricted shading, so this cannot use
        # openpyxl's read-only mode.
        wb = openpyxl.load_workbook(io.BytesIO(source), data_only=True)
        try:
            if sheet not in wb.sheetnames:
                raise ValueError(
                    f"Sheet '{sheet}' not found. Available: {wb.sheetnames}"
                )
            ws = wb[sheet]
            if ws.max_row < HEADER_ROW:
                raise ValueError("Master Planner sheet has no header row.")
            header_cells = list(ws[HEADER_ROW])
            body = list(ws.iter_rows(min_row=HEADER_ROW + 1, values_only=True))
        finally:
            wb.close()

        header_values = tuple(c.value for c in header_cells)
        col_index = MasterPlannerParser._locate_columns(header_values)
        reserved = set(col_index.values())

        target_year = (
            year if year is not None
            else MasterPlannerParser._pick_year(body, col_index)
        )
        rows = self._year_rows(body, col_index, target_year, horizon_weeks)

        columns, ignored = self._customer_columns(header_cells, reserved, rows)
        sites, warnings = derive_sites(columns, assign_stable_id, horizon_weeks)

        if target_year is None:
            warnings.append(
                "Could not determine which year to read from the manufacturing "
                "dates; all dated rows were used."
            )
        if not rows:
            warnings.append(
                "No schedule rows were found for the selected year, so every "
                "site came out inactive. Check the sheet and year."
            )

        return DerivedSiteSet(
            sites=sites,
            warnings=warnings,
            year=target_year,
            horizon_weeks=horizon_weeks,
            ignored_columns=ignored,
        )

    # ------------------------------------------------------------------
    # Extraction helpers
    # ------------------------------------------------------------------

    @staticmethod
    def _year_rows(
        body: list[tuple],
        col_index: dict[str, int],
        target_year: int | None,
        horizon_weeks: int,
    ) -> list[tuple[int, tuple]]:
        """Return ``(week_number, row)`` pairs for the selected year.

        The first row seen for a week wins, mirroring the Master Planner parser
        so the converted file and the comparison read the same weeks.
        """
        wk_col = col_index["weeks"]
        mfg_col = col_index.get("mfg")
        out: list[tuple[int, tuple]] = []
        seen: set[int] = set()

        for row in body:
            if wk_col >= len(row):
                continue
            week, ok = _to_int(row[wk_col])
            if not ok or not (1 <= week <= horizon_weeks) or week in seen:
                continue
            if target_year is not None:
                if mfg_col is None or mfg_col >= len(row):
                    continue
                mfg = row[mfg_col]
                if getattr(mfg, "year", None) != target_year:
                    continue
            seen.add(week)
            out.append((week, row))
        return out

    @staticmethod
    def _customer_columns(
        header_cells: list,
        reserved: set[int],
        rows: list[tuple[int, tuple]],
    ) -> tuple[list[PlannerColumn], list[str]]:
        """Collect the customer columns and their scheduled weeks."""
        columns: list[PlannerColumn] = []
        ignored: list[str] = []

        for idx, cell in enumerate(header_cells):
            if idx in reserved:
                continue
            if not _is_customer_column(cell.value):
                if cell.value is not None and str(cell.value).strip():
                    ignored.append(str(cell.value).strip())
                continue
            marks = tuple(
                week for week, row in rows
                if idx < len(row) and _is_schedule_mark(row[idx])
            )
            columns.append(
                PlannerColumn(
                    header=str(cell.value).strip(),
                    marks=marks,
                    eu_restricted=_fill_rgb(cell) == EU_RESTRICTED_FILL,
                    column_letter=get_column_letter(idx + 1),
                )
            )
        return columns, ignored

    # ------------------------------------------------------------------
    # Output
    # ------------------------------------------------------------------

    def write(self, result: DerivedSiteSet) -> bytes:
        """Render the conversion as a three-sheet workbook."""
        sheets = (
            (SITES_SHEET, sites_frame(result.sites)),
            (MAPPING_SHEET, mapping_frame(result.sites)),
            (NOTES_SHEET, notes_frame(result.sites)),
        )
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            for name, frame in sheets:
                frame.to_excel(writer, sheet_name=name, index=False)
            _autofit(writer)
        return buffer.getvalue()


def _autofit(writer, limit: int = 60) -> None:
    """Widen columns to roughly fit their contents, capped so headers stay sane."""
    for worksheet in writer.book.worksheets:
        for column_cells in worksheet.columns:
            longest = max(
                (len(str(c.value)) for c in column_cells if c.value is not None),
                default=0,
            )
            letter = get_column_letter(column_cells[0].column)
            worksheet.column_dimensions[letter].width = min(max(longest + 2, 10), limit)


def convert_master_planner(
    source: bytes,
    sheet: str = "Schedule",
    horizon_weeks: int = 52,
    year: int | None = None,
) -> DerivedSiteSet:
    """Module-level convenience wrapper around :class:`MasterPlannerConverter`."""
    return MasterPlannerConverter().convert(source, sheet, horizon_weeks, year)
