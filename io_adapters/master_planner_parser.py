"""Adapter: parse the wide Master Planner workbook.

The Master Planner ``Schedule`` sheet has one row per production week and one
column per customer, plus aggregate columns (``Total Commercial``, ``QC GEN``,
``US Demand``, ``RoW Demand``). Week numbers repeat across fiscal years, so a
year window must be selected.

Customer columns are matched to input-file ``Site_ID`` values by the leading
numeric token in the header (e.g. ``00449``). Columns without a leading number
receive a deterministic generated identifier (``RF-<hash8>``) so the planning
team can be given a stable mapping to use in future input sheets.

See .kiro/specs/optimizer-enhancements/design.md, Feature 1. Only this layer
touches openpyxl.
"""

from __future__ import annotations

import hashlib
import io
import re
from dataclasses import dataclass, field
from datetime import date, datetime
from typing import Any

import openpyxl

HEADER_ROW = 2

# Aggregate / non-customer headers that must never be treated as customers.
AGGREGATE_HEADERS = {
    "weeks", "mfg date holidays", "mfg date", "calibration date", "problem",
    "month", "fy", "us demand", "row demand", "total commercial", "qc gen",
    "sr82 req extra 1", "sr82 req", "strontium supplier", "po", "ordered",
}

# Header fragments identifying non-customer marker columns.
NON_CUSTOMER_FRAGMENTS = (
    "stab", "connector", "comment", "project", "testing",
    "asa", "blocked stock", "month coverage", "total inventory",
    "reception", "release", "sampling", "schedule",
)

_LEADING_NUM = re.compile(r"^(\d+)")


@dataclass
class AssignedId:
    """A generated stable identifier for an unnumbered customer column."""

    generated_id: str
    column_header: str
    customer_name: str
    normalized_key: str


@dataclass
class MasterPlannerData:
    """Structured contents of the Master Planner Schedule sheet."""

    weekly_planned_production: list[int] = field(default_factory=list)  # commercial + QC
    weekly_commercial: list[int] = field(default_factory=list)
    weekly_qc: list[int] = field(default_factory=list)
    customer_schedule: dict[str, list[int]] = field(default_factory=dict)
    assigned_ids: list[AssignedId] = field(default_factory=list)
    ignored_columns: list[str] = field(default_factory=list)
    week_to_row: dict[int, int] = field(default_factory=dict)
    mfg_dates: list[Any] = field(default_factory=list)   # 1-indexed, date | None
    cal_dates: list[Any] = field(default_factory=list)   # 1-indexed, date | None
    issues: list[str] = field(default_factory=list)
    rows_excluded: int = 0


def _normalize(text: str) -> str:
    """Lowercase, strip punctuation, collapse whitespace — the stable-ID key."""
    no_punct = re.sub(r"[^\w\s]", "", str(text).strip().lower())
    return re.sub(r"\s+", " ", no_punct).strip()


def assign_stable_id(header: str, digest_len: int = 8) -> str:
    """Return a deterministic ``RF-<hash>`` id derived from the column header.

    The same header always yields the same identifier, so re-parsing the same
    workbook (or the team reusing the shared id) is stable across runs.
    """
    key = _normalize(header)
    digest = hashlib.sha1(key.encode("utf-8")).hexdigest()[:digest_len]
    return f"RF-{digest}"


def _customer_name(header: str) -> str:
    """Strip a leading account number from the header to get the display name."""
    return _LEADING_NUM.sub("", str(header).strip()).strip()


def _is_customer_column(header: Any) -> bool:
    if header is None:
        return False
    text = str(header).strip()
    if not text:
        return False
    low = _normalize(text)
    if not low:
        return False
    # AGGREGATE_HEADERS are stored already-normalized (punctuation stripped).
    if any(low == a or low.startswith(a + " ") for a in AGGREGATE_HEADERS):
        return False
    if any(frag in low for frag in NON_CUSTOMER_FRAGMENTS):
        return False
    return True


def _to_date(value: Any):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def _is_schedule_mark(value: Any) -> bool:
    """True when a customer cell marks one scheduled generator.

    The Master Planner marks a scheduled generator with the integer 1. Booleans
    are excluded explicitly (``True == 1`` in Python).
    """
    if isinstance(value, bool):
        return False
    return isinstance(value, int) and value == 1


def _to_int(value: Any) -> tuple[int, bool]:
    """Return (int_value, ok). Non-numeric or negative yields (0, False)."""
    if value is None or value == "":
        return 0, True
    try:
        num = int(float(value))
    except (TypeError, ValueError):
        return 0, False
    if num < 0:
        return 0, False
    return num, True


class MasterPlannerParser:
    """Parses the Master Planner workbook. Implements ``MasterPlannerReaderPort``."""

    def parse(
        self,
        source: bytes,
        sheet: str = "Schedule",
        horizon_weeks: int = 52,
        year: int | None = None,
    ) -> MasterPlannerData:
        """Parse the Schedule sheet into a :class:`MasterPlannerData`.

        Parameters
        ----------
        source : bytes
            Raw workbook bytes.
        sheet : str
            Worksheet name (default ``Schedule``).
        horizon_weeks : int
            Planning horizon; weeks outside 1..horizon are excluded.
        year : int | None
            Which calendar year of rows to use. Week numbers repeat across years
            in this workbook. When ``None``, the most recent year present that
            has non-zero commercial demand is chosen.
        """
        wb = openpyxl.load_workbook(io.BytesIO(source), data_only=True, read_only=True)
        try:
            if sheet not in wb.sheetnames:
                raise ValueError(
                    f"Sheet '{sheet}' not found. Available: {wb.sheetnames}"
                )
            ws = wb[sheet]
            rows = list(ws.iter_rows(min_row=1, values_only=True))
        finally:
            wb.close()

        if len(rows) < HEADER_ROW:
            raise ValueError("Master Planner sheet has no header row.")

        header = rows[HEADER_ROW - 1]
        col_index = self._locate_columns(header)
        data = MasterPlannerData()

        # Identify customer columns and their identifiers
        customer_cols: dict[int, str] = {}
        for idx, raw_header in enumerate(header):
            if idx in col_index.values():
                continue
            if not _is_customer_column(raw_header):
                if raw_header is not None and str(raw_header).strip():
                    data.ignored_columns.append(str(raw_header).strip())
                continue
            text = str(raw_header).strip()
            match = _LEADING_NUM.match(text)
            if match:
                site_id = match.group(1)
            else:
                site_id = assign_stable_id(text)
                data.assigned_ids.append(
                    AssignedId(
                        generated_id=site_id,
                        column_header=text,
                        customer_name=_customer_name(text),
                        normalized_key=_normalize(text),
                    )
                )
            customer_cols[idx] = site_id

        # Choose the year window
        body = rows[HEADER_ROW:]
        target_year = year if year is not None else self._pick_year(body, col_index)

        # Initialize 1-indexed arrays
        T = horizon_weeks
        data.weekly_commercial = [0] * (T + 1)
        data.weekly_qc = [0] * (T + 1)
        data.weekly_planned_production = [0] * (T + 1)
        data.mfg_dates = [None] * (T + 1)
        data.cal_dates = [None] * (T + 1)
        for site_id in set(customer_cols.values()):
            data.customer_schedule.setdefault(site_id, [0] * (T + 1))

        wk_col = col_index["weeks"]
        seen_weeks: set[int] = set()

        for offset, row in enumerate(body):
            if wk_col >= len(row):
                continue
            week_val, ok = _to_int(row[wk_col])
            if not ok or week_val == 0:
                continue

            mfg = _to_date(row[col_index["mfg"]]) if col_index.get("mfg") is not None else None
            if target_year is not None:
                if mfg is None or mfg.year != target_year:
                    continue

            if not (1 <= week_val <= T):
                data.rows_excluded += 1
                continue
            if week_val in seen_weeks:
                continue  # first matching row for this week wins
            seen_weeks.add(week_val)

            data.week_to_row[week_val] = HEADER_ROW + offset + 1
            data.mfg_dates[week_val] = mfg
            if col_index.get("cal") is not None:
                data.cal_dates[week_val] = _to_date(row[col_index["cal"]])

            commercial, ok_c = _to_int(row[col_index["commercial"]]) if col_index.get("commercial") is not None else (0, True)
            if not ok_c:
                data.issues.append(
                    f"Week {week_val}: non-numeric or negative Total Commercial; treated as 0."
                )
            qc, ok_q = _to_int(row[col_index["qc"]]) if col_index.get("qc") is not None else (0, True)
            if not ok_q:
                data.issues.append(
                    f"Week {week_val}: non-numeric or negative QC GEN; treated as 0."
                )

            data.weekly_commercial[week_val] = commercial
            data.weekly_qc[week_val] = qc
            data.weekly_planned_production[week_val] = commercial + qc

            # Per-customer schedule marks.
            # A scheduled generator is recorded as the integer 1 in the
            # customer's cell. Verified against Total Commercial: the count of
            # cells equal to 1 matches the aggregate exactly. Other numeric
            # values in trailing columns are unrelated data and are ignored.
            for idx, site_id in customer_cols.items():
                if idx >= len(row):
                    continue
                if _is_schedule_mark(row[idx]):
                    data.customer_schedule[site_id][week_val] += 1

        return data

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------

    @staticmethod
    def _locate_columns(header: tuple) -> dict[str, int]:
        """Map logical names to 0-based column indexes by header text."""
        found: dict[str, int] = {}
        for idx, raw in enumerate(header):
            if raw is None:
                continue
            low = _normalize(raw)
            if low.startswith("weeks") and "weeks" not in found:
                found["weeks"] = idx
            elif low.startswith("mfg date") and "mfg" not in found:
                found["mfg"] = idx
            elif low.startswith("calibration date") and "cal" not in found:
                found["cal"] = idx
            elif low == "total commercial" and "commercial" not in found:
                found["commercial"] = idx
            elif low == "qc gen" and "qc" not in found:
                found["qc"] = idx
            elif low == "us demand" and "us" not in found:
                found["us"] = idx
            elif low == "row demand" and "row" not in found:
                found["row"] = idx
        if "weeks" not in found:
            raise ValueError("Could not locate the 'Weeks #' column in the header row.")
        return found

    @staticmethod
    def _pick_year(body: list[tuple], col_index: dict[str, int]) -> int | None:
        """Pick the most complete year of schedule data.

        Week numbers repeat per fiscal year in this workbook, so a year window
        must be chosen. We prefer the year with the most weeks carrying non-zero
        commercial demand (a fully-planned year), breaking ties toward the most
        recent year. This avoids selecting a partially-filled future year.
        """
        mfg_col = col_index.get("mfg")
        comm_col = col_index.get("commercial")
        if mfg_col is None:
            return None

        weeks_with_demand: dict[int, int] = {}
        any_rows: dict[int, int] = {}
        for row in body:
            if mfg_col >= len(row):
                continue
            d = _to_date(row[mfg_col])
            if d is None:
                continue
            any_rows[d.year] = any_rows.get(d.year, 0) + 1
            qty = 0
            if comm_col is not None and comm_col < len(row):
                qty, _ = _to_int(row[comm_col])
            if qty > 0:
                weeks_with_demand[d.year] = weeks_with_demand.get(d.year, 0) + 1

        if weeks_with_demand:
            best = max(weeks_with_demand.values())
            candidates = [y for y, c in weeks_with_demand.items() if c == best]
            return max(candidates)
        return max(any_rows) if any_rows else None


def parse_master_planner(
    source: bytes,
    sheet: str = "Schedule",
    horizon_weeks: int = 52,
    year: int | None = None,
) -> MasterPlannerData:
    """Module-level convenience wrapper around :class:`MasterPlannerParser`."""
    return MasterPlannerParser().parse(source, sheet, horizon_weeks, year)
